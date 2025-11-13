import os
import gradio as gr
from github import Github
import pandas as pd
import openpyxl
from openpyxl import load_workbook

def generate_vba_macro(task_description):
    """Generate VBA macro based on task description"""
    if not task_description or not task_description.strip():
        return "' Please enter a task description to generate VBA code."
    
    task_lower = task_description.lower()
    vba_code = []
    vba_code.append("Sub AutoMacro()")
    vba_code.append(f"    ' Task: {task_description}")
    vba_code.append("    Application.ScreenUpdating = False")
    vba_code.append("    Application.DisplayAlerts = False")
    vba_code.append("    On Error GoTo ErrorHandler")
    vba_code.append("")
    
    # Pattern matching for common Excel tasks
    if any(word in task_lower for word in ['format', 'color', 'highlight', 'style']):
        if 'header' in task_lower or 'first row' in task_lower:
            vba_code.append("    ' Format header row")
            vba_code.append("    With ActiveSheet.Rows(1)")
            vba_code.append("        .Font.Bold = True")
            vba_code.append("        .Interior.Color = RGB(68, 114, 196)")
            vba_code.append("        .Font.Color = RGB(255, 255, 255)")
            vba_code.append("    End With")
        else:
            vba_code.append("    ' Format selected range")
            vba_code.append("    With Selection")
            vba_code.append("        .Font.Bold = True")
            vba_code.append("        .Interior.Color = RGB(255, 242, 204)")
            vba_code.append("    End With")
    
    if any(word in task_lower for word in ['sum', 'total', 'add', 'calculate']):
        vba_code.append("    ' Calculate sum")
        vba_code.append("    Dim lastRow As Long")
        vba_code.append("    lastRow = ActiveSheet.Cells(ActiveSheet.Rows.Count, 1).End(xlUp).Row")
        vba_code.append("    ActiveSheet.Cells(lastRow + 1, 1).Formula = \"=SUM(A1:A\" & lastRow & \")\"")
    
    if any(word in task_lower for word in ['delete', 'remove', 'clear']):
        if 'row' in task_lower:
            vba_code.append("    ' Delete empty rows")
            vba_code.append("    Dim i As Long")
            vba_code.append("    For i = ActiveSheet.UsedRange.Rows.Count To 1 Step -1")
            vba_code.append("        If WorksheetFunction.CountA(ActiveSheet.Rows(i)) = 0 Then")
            vba_code.append("            ActiveSheet.Rows(i).Delete")
            vba_code.append("        End If")
            vba_code.append("    Next i")
        elif 'column' in task_lower:
            vba_code.append("    ' Delete empty columns")
            vba_code.append("    Dim j As Long")
            vba_code.append("    For j = ActiveSheet.UsedRange.Columns.Count To 1 Step -1")
            vba_code.append("        If WorksheetFunction.CountA(ActiveSheet.Columns(j)) = 0 Then")
            vba_code.append("            ActiveSheet.Columns(j).Delete")
            vba_code.append("        End If")
            vba_code.append("    Next j")
        else:
            vba_code.append("    ' Clear selected range")
            vba_code.append("    Selection.ClearContents")
    
    if any(word in task_lower for word in ['sort', 'order', 'arrange']):
        vba_code.append("    ' Sort data")
        vba_code.append("    Dim lastRow As Long, lastCol As Long")
        vba_code.append("    lastRow = ActiveSheet.Cells(ActiveSheet.Rows.Count, 1).End(xlUp).Row")
        vba_code.append("    lastCol = ActiveSheet.Cells(1, ActiveSheet.Columns.Count).End(xlToLeft).Column")
        vba_code.append("    ActiveSheet.Sort.SortFields.Clear")
        vba_code.append("    ActiveSheet.Sort.SortFields.Add Key:=Range(\"A1\"), SortOn:=xlSortOnValues, Order:=xlAscending")
        vba_code.append("    With ActiveSheet.Sort")
        vba_code.append("        .SetRange Range(Cells(1, 1), Cells(lastRow, lastCol))")
        vba_code.append("        .Header = xlYes")
        vba_code.append("        .Apply")
        vba_code.append("    End With")
    
    if any(word in task_lower for word in ['filter', 'autofilter']):
        vba_code.append("    ' Apply autofilter")
        vba_code.append("    Dim lastRow As Long, lastCol As Long")
        vba_code.append("    lastRow = ActiveSheet.Cells(ActiveSheet.Rows.Count, 1).End(xlUp).Row")
        vba_code.append("    lastCol = ActiveSheet.Cells(1, ActiveSheet.Columns.Count).End(xlToLeft).Column")
        vba_code.append("    ActiveSheet.Range(Cells(1, 1), Cells(lastRow, lastCol)).AutoFilter")
    
    if any(word in task_lower for word in ['copy', 'duplicate']):
        vba_code.append("    ' Copy selected range")
        vba_code.append("    Selection.Copy")
        vba_code.append("    ' Paste to new location (modify as needed)")
        vba_code.append("    ' ActiveSheet.Range(\"A10\").PasteSpecial Paste:=xlPasteValues")
    
    if any(word in task_lower for word in ['save', 'export']):
        vba_code.append("    ' Save workbook")
        vba_code.append("    ActiveWorkbook.Save")
    
    if any(word in task_lower for word in ['print', 'printout']):
        vba_code.append("    ' Print active sheet")
        vba_code.append("    ActiveSheet.PrintOut")
    
    # Default macro if no specific pattern matched
    if len(vba_code) == 6:  # Only header code added
        vba_code.append("    ' Custom task - modify as needed")
        vba_code.append("    MsgBox \"Task completed: " + task_description + "\"")
    
    vba_code.append("")
    vba_code.append("    Application.ScreenUpdating = True")
    vba_code.append("    Application.DisplayAlerts = True")
    vba_code.append("    Exit Sub")
    vba_code.append("")
    vba_code.append("ErrorHandler:")
    vba_code.append("    Application.ScreenUpdating = True")
    vba_code.append("    Application.DisplayAlerts = True")
    vba_code.append("    MsgBox \"Error: \" & Err.Description")
    vba_code.append("End Sub")
    
    return "\n".join(vba_code)

def handle_excel_upload(file):
    """Analyze uploaded Excel file and return information"""
    if file is None:
        return "No file uploaded."
    
    try:
        file_path = file.name
        file_ext = os.path.splitext(file_path)[1].lower()
        
        if file_ext not in ['.xlsx', '.xls']:
            return f"Error: Unsupported file type {file_ext}. Please upload .xlsx or .xls files."
        
        # Try to read with openpyxl for .xlsx
        if file_ext == '.xlsx':
            wb = load_workbook(file_path, read_only=True)
            sheet_names = wb.sheetnames
            info = [f"✓ File '{os.path.basename(file_path)}' uploaded successfully."]
            info.append(f"📊 Sheets found: {len(sheet_names)}")
            
            for i, sheet_name in enumerate(sheet_names[:5], 1):  # Show first 5 sheets
                ws = wb[sheet_name]
                max_row = ws.max_row
                max_col = ws.max_column
                info.append(f"  Sheet {i}: '{sheet_name}' ({max_row} rows × {max_col} columns)")
            
            if len(sheet_names) > 5:
                info.append(f"  ... and {len(sheet_names) - 5} more sheets")
            
            wb.close()
            return "\n".join(info)
        else:
            # For .xls files, use pandas
            try:
                xls_file = pd.ExcelFile(file_path)
                sheet_names = xls_file.sheet_names
                info = [f"✓ File '{os.path.basename(file_path)}' uploaded successfully."]
                info.append(f"📊 Sheets found: {len(sheet_names)}")
                
                for i, sheet_name in enumerate(sheet_names[:5], 1):
                    df = pd.read_excel(xls_file, sheet_name=sheet_name, nrows=0)
                    info.append(f"  Sheet {i}: '{sheet_name}' ({len(df.columns)} columns)")
                
                if len(sheet_names) > 5:
                    info.append(f"  ... and {len(sheet_names) - 5} more sheets")
                
                return "\n".join(info)
            except Exception as e:
                return f"✓ File uploaded, but analysis failed: {str(e)}"
    
    except Exception as e:
        return f"Error processing file: {str(e)}"

# Initialize GitHub client
github_token = os.getenv("GITHUB_TOKEN")
if github_token and github_token != "your_github_token":
    try:
        g = Github(github_token)
        github_available = True
    except:
        github_available = False
else:
    github_available = False

def push_to_github(repo_name, file_name, code):
    """Push generated VBA code to GitHub repository"""
    if not github_available:
        return "⚠️ GitHub token not configured. Set GITHUB_TOKEN environment variable."
    
    if not repo_name or not repo_name.strip():
        return "⚠️ Please enter a repository name."
    
    if not file_name or not file_name.strip():
        return "⚠️ Please enter a file name."
    
    if not code or not code.strip():
        return "⚠️ No VBA code to push. Generate a macro first."
    
    # Ensure file has .bas extension for VBA modules
    if not file_name.endswith('.bas'):
        file_name = file_name + '.bas'
    
    try:
        repo = g.get_user().get_repo(repo_name)
        # Check if file already exists
        try:
            contents = repo.get_contents(file_name)
            # File exists, update it
            repo.update_file(file_name, "Update VBA macro", code, contents.sha)
            return f"✓ Successfully updated {file_name} in {repo_name}"
        except:
            # File doesn't exist, create it
            repo.create_file(file_name, "Add VBA macro", code)
            return f"✓ Successfully pushed {file_name} to {repo_name}"
    except Exception as e:
        error_msg = str(e)
        if "404" in error_msg or "Not Found" in error_msg:
            return f"⚠️ Repository '{repo_name}' not found. Please check the repository name and ensure it exists."
        elif "401" in error_msg or "Bad credentials" in error_msg:
            return "⚠️ GitHub authentication failed. Please check your GITHUB_TOKEN."
        else:
            return f"⚠️ GitHub push failed: {error_msg}"

# Create Gradio interface
with gr.Blocks(theme=gr.themes.Soft()) as demo:
    gr.Markdown("""
    # 🤖 ExcelBot Pro - VBA Automation Chatbot
    
    Automate Excel tasks by generating VBA macros from natural language descriptions.
    Upload Excel files, generate VBA code, and push to GitHub repositories.
    """)
    
    with gr.Tab("📝 Generate VBA Macro"):
        gr.Markdown("### Describe your Excel task in natural language")
        task_input = gr.Textbox(
            label="Task Description",
            placeholder="e.g., Format the header row with blue background, Sort data by column A, Delete empty rows...",
            lines=3
        )
        generate_button = gr.Button("Generate VBA Macro", variant="primary")
        vba_output = gr.Code(
            label="Generated VBA Macro",
            language="vbnet",
            lines=20
        )
    
    with gr.Tab("📤 Upload Excel File"):
        gr.Markdown("### Upload and analyze your Excel file")
        excel_file = gr.File(
            label="Upload Excel File (.xls/.xlsx)",
            file_types=[".xlsx", ".xls"]
        )
        upload_status = gr.Textbox(
            label="File Analysis",
            lines=8,
            interactive=False
        )
    
    with gr.Tab("🚀 Push to GitHub"):
        gr.Markdown("### Push generated VBA code to GitHub")
        if not github_available:
            gr.Markdown("⚠️ **GitHub integration not available.** Set `GITHUB_TOKEN` environment variable to enable.")
        
        with gr.Row():
            repo_name = gr.Textbox(
                label="Repository Name",
                placeholder="my-excel-macros"
            )
            file_name = gr.Textbox(
                label="File Name",
                placeholder="MyMacro.bas",
                value="AutoMacro.bas"
            )
        push_button = gr.Button("Push to GitHub", variant="primary")
        push_status = gr.Textbox(
            label="GitHub Status",
            lines=3,
            interactive=False
        )
    
    # Event handlers
    generate_button.click(
        fn=generate_vba_macro,
        inputs=[task_input],
        outputs=[vba_output]
    )
    
    excel_file.change(
        fn=handle_excel_upload,
        inputs=[excel_file],
        outputs=[upload_status]
    )
    
    push_button.click(
        fn=push_to_github,
        inputs=[repo_name, file_name, vba_output],
        outputs=[push_status]
    )

if __name__ == "__main__":
    demo.launch(server_name="0.0.0.0", server_port=7860)
